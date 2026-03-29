#!/usr/bin/env python3
# build_json.py
# Excel（id中心） -> data/gallery.json を生成する
from __future__ import annotations

import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "gallery.xlsx"          # Excelファイル名
SHEET_NAME = "gallery"                     # シート名
OUT_PATH = ROOT / "data" / "gallery.json" # 出力先
IMAGES_DIR = ROOT / "images"               # 画像フォルダ

# 許容する画像拡張子（優先順）
IMG_EXTS = [".png", ".jpg", ".jpeg", ".webp"]


def normalize_date(v) -> str:
    if v is None or str(v).strip() == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip().replace("/", "-")
    return s


def find_image_file(id_str: str) -> str | None:
    """
    images/ の中から id に一致するファイルを探す。
    例: id=0001 -> images/0001.png (無ければ jpg/jpeg/webp を探す)
    戻り値は HTML/JS 用の相対パス（images/xxxx.ext）
    """
    for ext in IMG_EXTS:
        p = IMAGES_DIR / f"{id_str}{ext}"
        if p.exists():
            return f"images/{p.name}"
    return None


def get_cell(row, idx_map, key):
    if key not in idx_map:
        return None
    i = idx_map[key]
    return row[i] if i < len(row) else None


def clean_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def first_non_empty(*values) -> str:
    for v in values:
        s = clean_text(v)
        if s:
            return s
    return ""


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"ERROR: Excelが見つかりません: {XLSX_PATH}")
        return 1

    if not IMAGES_DIR.exists():
        print(f"ERROR: imagesフォルダが見つかりません: {IMAGES_DIR}")
        return 1

    wb = load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: シート '{SHEET_NAME}' がありません。存在: {wb.sheetnames}")
        return 1

    ws = wb[SHEET_NAME]

    # 1行目: ヘッダー
    headers = []
    for c in ws[1]:
        headers.append(str(c.value).strip() if c.value is not None else "")

    if "id" not in headers:
        print("ERROR: ヘッダー行(1行目)に 'id' 列が必要です。")
        print(f"現在のヘッダー: {headers}")
        return 1

    idx = {h: i for i, h in enumerate(headers) if h}

    items: list[dict] = []
    warnings = 0

    for r in ws.iter_rows(min_row=2, values_only=True):
        rid = get_cell(r, idx, "id")
        if rid is None or str(rid).strip() == "":
            continue

        id_str = str(rid).strip()

        # fileは自動生成（images内を探索）
        file_path = find_image_file(id_str)
        if not file_path:
            warnings += 1
            print(f"WARNING: images/{id_str}.[png/jpg/jpeg/webp] が見つかりません（スキップ）")
            continue

        # 新構成を優先しつつ、旧構成(title)にも対応
        title_ja = first_non_empty(
            get_cell(r, idx, "title_ja"),
            get_cell(r, idx, "title")
        )
        title_en = first_non_empty(get_cell(r, idx, "title_en"))

        date = get_cell(r, idx, "date")

        # desc は現状そのまま日本語扱い
        desc_ja = first_non_empty(get_cell(r, idx, "desc"))
        desc_en = first_non_empty(get_cell(r, idx, "desc_en"))

        it = {
            "id": id_str,
            "title_en": title_en if title_en else f"Tokyo Neon Dystopia - {id_str}",
            "title_ja": title_ja,
            "file": file_path,
            "date": normalize_date(date),
            "desc": desc_ja,
            "desc_en": desc_en,
        }

        items.append(it)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {"items": items}
    OUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    print(f"OK: {OUT_PATH} を生成しました（{len(items)}件）")
    if warnings:
        print(f"NOTE: 画像未検出でスキップ {warnings}件（imagesフォルダとidを確認）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())